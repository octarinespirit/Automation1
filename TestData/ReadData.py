import openpyxl

vk = openpyxl.load_workbook("C://Automation1/TestData/TestData.xlsx")

def fetch_number_of_rows(Sheetname):
    sh = vk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell):
    sh = vk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value

print(fetch_number_of_rows("Sheet1"))
print(fetch_cell_data("Sheet1",1,1))

sh = vk['Sheet1']
print(sh.max_row)
cell = sh.cell(1,1)
print(cell.value)